import os
import uuid
import io
import re
import traceback
import xml.etree.ElementTree as ET

import pandas as pd

from flask import (
    Flask,
    request,
    send_file,
    after_this_request,
    render_template
)

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill


# ============================================================
# FLASK APP
# ============================================================

app = Flask(__name__)

UPLOAD_FOLDER = os.path.join(os.getcwd(), "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


# ============================================================
# FILE VALIDATION
# ============================================================

def allowed_file(filename):
    """
    Allow Excel and CSV files.
    """
    return (
        "." in filename
        and filename.rsplit(".", 1)[1].lower()
        in {"xls", "xlsx", "csv"}
    )


# ============================================================
# COLUMN NORMALIZATION
# ============================================================

def normalize_columns(df):
    """
    Normalize column names from different exports.
    """

    df = df.copy()

    df.columns = (
        pd.Index(df.columns)
        .astype(str)
        .str.replace("\xa0", " ", regex=False)
        .str.strip()
        .str.upper()
        .str.replace(r"\s+", " ", regex=True)
    )

    rename_map = {
        # Item name variations
        "ITEM": "ITEM NAME",
        "ITEMNAME": "ITEM NAME",
        "PRODUCT": "ITEM NAME",
        "PRODUCT NAME": "ITEM NAME",
        "NAME": "ITEM NAME",
        "ITEM DESCRIPTION": "ITEM NAME",
        "DESCRIPTION": "ITEM NAME",

        # Quantity variations
        "QUANTITY": "QTY",
        "SALE QTY": "QTY",
        "SALES QTY": "QTY",
        "SOLD QTY": "QTY",

        # Available stock variations
        "AVAILABLE QTY": "AVAIL. QTY",
        "AVAILABLE QUANTITY": "AVAIL. QTY",
        "AVAIL QTY": "AVAIL. QTY",
        "AVAIL QTY.": "AVAIL. QTY",
        "AVAILABLE STOCK": "AVAIL. QTY",

        # SKU variations
        "ITEM CODE": "SKU",
        "PRODUCT CODE": "SKU",
        "CODE": "SKU",
        "ITEM SKU": "SKU",
    }

    return df.rename(columns=rename_map)


# ============================================================
# SKU NORMALIZATION
# ============================================================

def normalize_sku(value):
    """
    Normalize SKU values to improve matching between
    sales and stock exports.

    Examples:

    A05CR       -> A05CR
    A05CR.0     -> A05CR
    A05 CR      -> A05CR
    A05-CR      -> A05CR
    A05_CR      -> A05CR
    A05CR NBSP  -> A05CR
    """

    if pd.isna(value):
        return ""

    sku = str(value)

    # Replace non-breaking and invisible spaces
    sku = sku.replace("\xa0", "")
    sku = sku.replace("\u200b", "")
    sku = sku.replace("\ufeff", "")

    sku = sku.strip().upper()

    # Remove all normal spaces
    sku = re.sub(r"\s+", "", sku)

    # Remove Excel numeric suffix
    # Example: A05CR.0 -> A05CR
    sku = re.sub(r"\.0+$", "", sku)

    # Remove common separators
    sku = sku.replace("-", "")
    sku = sku.replace("_", "")

    # Remove trailing punctuation
    sku = sku.strip(".,;:")

    return sku


# ============================================================
# DETECT HEADER ROW
# ============================================================

def detect_header(df):
    """
    Detect the real header row.

    The sales file should contain an item-related
    column and a quantity-related column.

    The stock file can contain SKU and quantity columns.
    """

    df = df.copy()

    search_limit = min(20, len(df))

    for i in range(search_limit):

        row = (
            df.iloc[i]
            .fillna("")
            .astype(str)
            .str.replace("\xa0", " ", regex=False)
            .str.strip()
            .str.upper()
            .tolist()
        )

        has_item = any(
            (
                "ITEM" in value
                or "PRODUCT" in value
                or value == "NAME"
            )
            for value in row
        )

        has_sku = any(
            (
                value == "SKU"
                or "ITEM CODE" in value
                or "PRODUCT CODE" in value
            )
            for value in row
        )

        has_qty = any(
            (
                "QTY" in value
                or "QUANTITY" in value
            )
            for value in row
        )

        # Sales header
        if has_item and has_qty:

            df.columns = (
                df.iloc[i]
                .fillna("")
                .astype(str)
            )

            df = df.iloc[i + 1:].copy()

            df.reset_index(
                drop=True,
                inplace=True
            )

            return df

        # Stock header
        if has_sku and has_qty:

            df.columns = (
                df.iloc[i]
                .fillna("")
                .astype(str)
            )

            df = df.iloc[i + 1:].copy()

            df.reset_index(
                drop=True,
                inplace=True
            )

            return df

    return df


# ============================================================
# EXCEL 2003 XML READER
# ============================================================

def read_excel_xml(bio):
    """
    Read Excel 2003 XML Spreadsheet files.

    Some files use .xls extensions but contain XML.

    Supports skipped cells through ss:Index.
    """

    bio.seek(0)

    tree = ET.parse(bio)
    root = tree.getroot()

    namespace_uri = (
        "urn:schemas-microsoft-com:office:spreadsheet"
    )

    ns = {
        "ss": namespace_uri
    }

    rows = []

    worksheets = root.findall(
        ".//ss:Worksheet",
        ns
    )

    if not worksheets:
        raise ValueError(
            "No worksheet found in Excel XML file."
        )

    # Use first worksheet
    worksheet = worksheets[0]

    table = worksheet.find(
        "ss:Table",
        ns
    )

    if table is None:
        raise ValueError(
            "No table found in Excel XML file."
        )

    for row in table.findall(
        "ss:Row",
        ns
    ):

        row_data = []

        for cell in row.findall(
            "ss:Cell",
            ns
        ):

            # Handle skipped columns
            index = cell.get(
                f"{{{namespace_uri}}}Index"
            )

            if index:

                index = int(index)

                while len(row_data) < index - 1:
                    row_data.append("")

            data = cell.find(
                "ss:Data",
                ns
            )

            value = (
                data.text
                if (
                    data is not None
                    and data.text is not None
                )
                else ""
            )

            row_data.append(value)

        rows.append(row_data)

    if not rows:
        raise ValueError(
            "No data rows found in Excel XML file."
        )

    # Ensure all rows have the same length
    max_columns = max(
        len(row)
        for row in rows
    )

    normalized_rows = []

    for row in rows:

        row = row + (
            [""] * (
                max_columns - len(row)
            )
        )

        normalized_rows.append(
            row
        )

    return pd.DataFrame(
        normalized_rows
    )


# ============================================================
# MAIN FILE READER
# ============================================================

def read_file(
    file_storage,
    header_row=None
):
    """
    Read XLSX, XLS, Excel XML or CSV.
    """

    content = file_storage.read()

    file_storage.seek(0)

    if not content:
        raise ValueError(
            "Empty file uploaded."
        )

    bio = io.BytesIO(content)

    # --------------------------------------------------------
    # XLSX
    # --------------------------------------------------------

    if content[:2] == b"PK":

        bio.seek(0)

        return pd.read_excel(
            bio,
            engine="openpyxl",
            header=header_row,
            dtype=str
        )

    # --------------------------------------------------------
    # TRADITIONAL XLS
    # --------------------------------------------------------

    if content[:4] == b"\xD0\xCF\x11\xE0":

        bio.seek(0)

        return pd.read_excel(
            bio,
            engine="xlrd",
            header=header_row,
            dtype=str
        )

    # --------------------------------------------------------
    # EXCEL 2003 XML
    # --------------------------------------------------------

    stripped = content.lstrip()

    if (
        stripped.startswith(b"<?xml")
        or b"<Workbook" in content[:2000]
        or b":Workbook" in content[:2000]
    ):

        bio.seek(0)

        return read_excel_xml(
            bio
        )

    # --------------------------------------------------------
    # CSV
    # --------------------------------------------------------

    bio.seek(0)

    return pd.read_csv(
        bio,
        sep=None,
        engine="python",
        header=header_row,
        encoding_errors="ignore",
        dtype=str,
        on_bad_lines="skip",
        quoting=3
    )


# ============================================================
# EXTRACT SALES SKU
# ============================================================

def extract_sales_sku(item_name):
    """
    Extract SKU from the first word of the item name.

    Example:

    A05CR SAMSUNG A05 COMP LCD

    Returns:

    A05CR
    """

    if pd.isna(item_name):
        return ""

    item_name = str(item_name).strip()

    if not item_name:
        return ""

    first_word = item_name.split()[0]

    return normalize_sku(
        first_word
    )


# ============================================================
# FORMAT EXCEL WORKSHEET
# ============================================================

def format_worksheet(
    ws,
    item_name_width=100
):
    """
    Apply common formatting to a worksheet.
    """

    # Freeze first row
    ws.freeze_panes = "A2"

    # Enable filters
    ws.auto_filter.ref = ws.dimensions

    # Header formatting
    for cell in ws[1]:

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # Number formatting
    headers = {
        cell.column: str(cell.value)
        for cell in ws[1]
    }

    numeric_headers = {
        "QTY",
        "CURRENT_STOCK",
        "REORDER_QTY"
    }

    for row in range(
        2,
        ws.max_row + 1
    ):

        for column_index, header in headers.items():

            if header in numeric_headers:

                ws.cell(
                    row=row,
                    column=column_index
                ).number_format = "#,##0"

    # Column widths
    for column_cells in ws.columns:

        column_letter = (
            column_cells[0]
            .column_letter
        )

        header = str(
            column_cells[0].value
        )

        if header == "ITEM NAME":

            ws.column_dimensions[
                column_letter
            ].width = item_name_width

            continue

        if header in {
            "SKU",
            "ORIGINAL_SKU"
        }:

            ws.column_dimensions[
                column_letter
            ].width = 20

            continue

        if header == "STOCK_MATCH":

            ws.column_dimensions[
                column_letter
            ].width = 18

            continue

        max_length = 0

        for cell in column_cells:

            value = (
                ""
                if cell.value is None
                else str(cell.value)
            )

            max_length = max(
                max_length,
                len(value)
            )

        ws.column_dimensions[
            column_letter
        ].width = min(
            max_length + 4,
            40
        )


# ============================================================
# INDEX
# ============================================================

@app.route("/")
def index():

    return render_template(
        "index.html"
    )


# ============================================================
# PROCESS FILES
# ============================================================

@app.route(
    "/process",
    methods=["POST"]
)
def process():

    try:

        # ----------------------------------------------------
        # VALIDATE UPLOADS
        # ----------------------------------------------------

        if (
            "sales" not in request.files
            or "stock" not in request.files
        ):

            return (
                "Upload both sales and stock files.",
                400
            )

        sales_file = request.files[
            "sales"
        ]

        stock_file = request.files[
            "stock"
        ]

        if not sales_file.filename:

            return (
                "Please select a sales file.",
                400
            )

        if not stock_file.filename:

            return (
                "Please select a stock file.",
                400
            )

        if not allowed_file(
            sales_file.filename
        ):

            return (
                "Sales file must be XLS, XLSX or CSV.",
                400
            )

        if not allowed_file(
            stock_file.filename
        ):

            return (
                "Stock file must be XLS, XLSX or CSV.",
                400
            )

        # ----------------------------------------------------
        # READ RAW FILES
        # ----------------------------------------------------

        df_sales_raw = read_file(
            sales_file,
            header_row=None
        )

        df_stock_raw = read_file(
            stock_file,
            header_row=None
        )

        # ----------------------------------------------------
        # DETECT HEADERS
        # ----------------------------------------------------

        df_sales = detect_header(
            df_sales_raw
        )

        df_stock = detect_header(
            df_stock_raw
        )

        # ----------------------------------------------------
        # NORMALIZE COLUMN NAMES
        # ----------------------------------------------------

        df_sales = normalize_columns(
            df_sales
        )

        df_stock = normalize_columns(
            df_stock
        )

        # ----------------------------------------------------
        # VALIDATE SALES COLUMNS
        # ----------------------------------------------------

        required_sales = {
            "ITEM NAME",
            "QTY"
        }

        missing_sales = (
            required_sales
            - set(df_sales.columns)
        )

        if missing_sales:

            return (
                "Sales file missing required columns: "
                f"{', '.join(sorted(missing_sales))}. "
                f"Found columns: {list(df_sales.columns)}",
                400
            )

        # ----------------------------------------------------
        # VALIDATE STOCK COLUMNS
        # ----------------------------------------------------

        required_stock = {
            "SKU",
            "ITEM NAME",
            "AVAIL. QTY"
        }

        missing_stock = (
            required_stock
            - set(df_stock.columns)
        )

        if missing_stock:

            return (
                "Stock file missing required columns: "
                f"{', '.join(sorted(missing_stock))}. "
                f"Found columns: {list(df_stock.columns)}",
                400
            )

        # ====================================================
        # CLEAN SALES
        # ====================================================

        df_sales = df_sales.copy()

        df_sales["ITEM NAME"] = (
            df_sales["ITEM NAME"]
            .fillna("")
            .astype(str)
            .str.replace("\xa0", " ", regex=False)
            .str.strip()
        )

        # Remove empty item names
        df_sales = df_sales[
            df_sales["ITEM NAME"] != ""
        ]

        # Preserve original extracted SKU
        df_sales["ORIGINAL_SKU"] = (
            df_sales["ITEM NAME"]
            .apply(
                lambda value:
                str(value).split()[0]
                if str(value).strip()
                else ""
            )
        )

        # Normalize SKU
        df_sales["SKU"] = (
            df_sales["ORIGINAL_SKU"]
            .apply(normalize_sku)
        )

        # Convert quantity
        df_sales["QTY"] = pd.to_numeric(
            df_sales["QTY"],
            errors="coerce"
        ).fillna(0)

        # Remove invalid SKUs
        df_sales = df_sales[
            df_sales["SKU"].notna()
            & (df_sales["SKU"] != "")
            & (df_sales["SKU"] != "NAN")
            & (df_sales["SKU"] != "NONE")
        ]

        # ====================================================
        # CLEAN STOCK
        # ====================================================

        df_stock = df_stock.copy()

        # Preserve original stock SKU
        df_stock["ORIGINAL_SKU"] = (
            df_stock["SKU"]
            .fillna("")
            .astype(str)
        )

        # Normalize stock SKU
        df_stock["SKU"] = (
            df_stock["ORIGINAL_SKU"]
            .apply(normalize_sku)
        )

        # Clean stock item name
        df_stock["ITEM NAME"] = (
            df_stock["ITEM NAME"]
            .fillna("")
            .astype(str)
            .str.replace("\xa0", " ", regex=False)
            .str.strip()
        )

        # Convert stock quantity
        df_stock["AVAIL. QTY"] = pd.to_numeric(
            df_stock["AVAIL. QTY"],
            errors="coerce"
        ).fillna(0)

        # Remove invalid SKU rows
        df_stock = df_stock[
            df_stock["SKU"].notna()
            & (df_stock["SKU"] != "")
            & (df_stock["SKU"] != "NAN")
            & (df_stock["SKU"] != "NONE")
        ]

        # ====================================================
        # AGGREGATE STOCK BY SKU
        # ====================================================

        df_stock_agg = (
            df_stock
            .groupby(
                "SKU",
                as_index=False
            )
            .agg(
                CURRENT_STOCK=(
                    "AVAIL. QTY",
                    "sum"
                )
            )
        )

        # ====================================================
        # AGGREGATE SALES
        # ====================================================

        sales_summary = (
            df_sales
            .groupby(
                [
                    "SKU",
                    "ITEM NAME"
                ],
                as_index=False
            )
            .agg(
                QTY=(
                    "QTY",
                    "sum"
                )
            )
        )

        # ====================================================
        # EXACT SKU MATCH
        # ====================================================

        merged = pd.merge(
            sales_summary,
            df_stock_agg,
            on="SKU",
            how="left",
            indicator=True
        )

        # Match status
        merged["STOCK_MATCH"] = (
            merged["_merge"]
            .map(
                {
                    "both": "MATCHED",
                    "left_only": "NOT FOUND"
                }
            )
        )

        merged.drop(
            columns=["_merge"],
            inplace=True
        )

        # Unmatched stock is displayed as zero
        merged["CURRENT_STOCK"] = (
            merged["CURRENT_STOCK"]
            .fillna(0)
        )

        # ====================================================
        # CALCULATE REORDER QUANTITY
        # ====================================================

        merged["REORDER_QTY"] = (
            merged["QTY"]
            - merged["CURRENT_STOCK"]
        ).clip(
            lower=0
        )

        # ====================================================
        # FINAL COLUMN ORDER
        # ====================================================

        merged = merged[
            [
                "SKU",
                "ITEM NAME",
                "QTY",
                "CURRENT_STOCK",
                "REORDER_QTY",
                "STOCK_MATCH"
            ]
        ]

        # Sort by sales quantity
        merged = merged.sort_values(
            by="QTY",
            ascending=False
        )

        merged.reset_index(
            drop=True,
            inplace=True
        )

        # ====================================================
        # SPLIT MATCHED AND UNMATCHED
        # ====================================================

        matched_items = merged[
            merged["STOCK_MATCH"]
            == "MATCHED"
        ].copy()

        unmatched_items = merged[
            merged["STOCK_MATCH"]
            == "NOT FOUND"
        ].copy()

        # ====================================================
        # CREATE OUTPUT FILE
        # ====================================================

        filename = (
            f"reorder_"
            f"{uuid.uuid4().hex[:8]}"
            f".xlsx"
        )

        path = os.path.join(
            UPLOAD_FOLDER,
            filename
        )

        # ====================================================
        # WRITE EXCEL REPORT
        # ====================================================

        with pd.ExcelWriter(
            path,
            engine="openpyxl"
        ) as writer:

            # Main report
            merged.to_excel(
                writer,
                sheet_name="REORDER REPORT",
                index=False
            )

            # Matched items
            matched_items.to_excel(
                writer,
                sheet_name="MATCHED ITEMS",
                index=False
            )

            # Unmatched items
            unmatched_items.to_excel(
                writer,
                sheet_name="UNMATCHED SKUS",
                index=False
            )

        # ====================================================
        # FORMAT WORKBOOK
        # ====================================================

        wb = load_workbook(
            path
        )

        for sheet_name in wb.sheetnames:

            ws = wb[
                sheet_name
            ]

            format_worksheet(
                ws
            )

        wb.save(
            path
        )

        # ====================================================
        # CLEANUP FILE AFTER DOWNLOAD
        # ====================================================

        @after_this_request
        def cleanup(response):

            try:

                if os.path.exists(
                    path
                ):

                    os.remove(
                        path
                    )

            except Exception as error:

                print(
                    "File cleanup error:",
                    error
                )

            return response

        # ====================================================
        # DOWNLOAD
        # ====================================================

        return send_file(
            path,
            as_attachment=True,
            download_name="reorder_report.xlsx"
        )

    except Exception:

        error_trace = (
            traceback.format_exc()
        )

        print(
            error_trace
        )

        return (
            f"<pre>{error_trace}</pre>",
            500
        )


# ============================================================
# RUN APPLICATION
# ============================================================

if __name__ == "__main__":

    port = int(
        os.environ.get(
            "PORT",
            5000
        )
    )

    app.run(
        host="0.0.0.0",
        port=port,
        debug=True
    )